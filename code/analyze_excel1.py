import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "FileCarving_Results_final.xlsx")
CHARTS_DIR = os.path.join(BASE_DIR, "charts")

TOOLS = ["Foremost", "PhotoRec", "Scalpel"]
SCENARIOS = ["Scenario 1 (Normal)", "Scenario 2 (Fragmented)"]

# File types as they appear in the original dataset (lowercase)
FILE_TYPES_RAW = ["jpg", "png", "pdf", "docx", "mp4"]
FILE_TYPES_LABEL = ["JPG", "PNG", "PDF", "DOCX", "MP4"]

DATASET_SHEET = "1 - Original Dataset"
TOOL_SHEETS = {
    "Foremost": {
        "Scenario 1 (Normal)": "2 - Foremost S1",
        "Scenario 2 (Fragmented)": "3 - Foremost S2",
    },
    "PhotoRec": {
        "Scenario 1 (Normal)": "4 - PhotoRec S1",
        "Scenario 2 (Fragmented)": "5 - PhotoRec S2",
    },
    "Scalpel": {
        "Scenario 1 (Normal)": "6 - Scalpel S1",
        "Scenario 2 (Fragmented)": "7 - Scalpel S2",
    },
}
SUMMARY_SHEET = "8 - Results Summary"
FILETYPE_SHEET = "9 - Per File Type"

# Colors
HEADER_BLUE = "1F4E79"
WHITE = "FFFFFF"
GREEN = "E2EFDA"
YELLOW = "FFF2CC"
LIGHT_BLUE = "DEEAF1"

SCENARIO_COLORS = ["#2196F3", "#FF7043"]
TOOL_COLORS = {"Foremost": "#2196F3", "PhotoRec": "#4CAF50", "Scalpel": "#FF9800"}
FILE_TYPE_COLORS = ["#5C9BD6", "#ED7D31", "#A9D18E", "#FFC000", "#9E48B5"]


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def hcell(cell, text=None, bg=WHITE, bold=False, center=True, num_fmt=None, font_color="000000"):
    if text is not None:
        cell.value = text
    cell.font = Font(bold=bold, size=10, name="Arial", color=font_color)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    if num_fmt:
        cell.number_format = num_fmt


def save_chart(fig, filename):
    os.makedirs(CHARTS_DIR, exist_ok=True)
    path = os.path.join(CHARTS_DIR, filename)
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"    [+] Saved: {filename}")


# ─────────────────────────────────────────────
# READ DATA
# ─────────────────────────────────────────────
def read_original_dataset(wb):
    ws = wb[DATASET_SHEET]
    dataset = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 5:
            continue
        name = str(row[1]).strip() if row[1] else None
        ftype = str(row[2]).strip().lower() if row[2] else None
        sha = str(row[4]).strip().lower() if row[4] else None
        if name and ftype and sha:
            dataset[sha] = {"type": ftype, "name": name}
    print(f"    Dataset: {len(dataset)} files")
    return dataset


def read_tool_output(wb, sheet_name, dataset):
    ws = wb[sheet_name]
    recovered = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[1]:
            continue
        name = str(row[1]).strip() if row[1] else ""
        ftype = str(row[2]).strip().lower() if row[2] else "other"
        sha = str(row[4]).strip().lower() if row[4] else ""
        status_raw = str(row[6]).strip().lower() if row[6] else ""

        if sha in dataset:
            status = "Valid"
        elif "corrupt" in status_raw:
            status = "Corrupted"
        else:
            status = "False Positive"

        # Only add rows that look like file entries (not summary rows)
        if name and not name.startswith("Total") and not name.startswith("Valid") and not name.startswith("Corrupted") and not name.startswith("False"):
            recovered.append({"name": name, "type": ftype, "sha": sha, "status": status})
    return recovered


# ─────────────────────────────────────────────
# METRICS
# ─────────────────────────────────────────────
def compute_metrics(recovered, dataset, proc_time=""):
    rel = len(dataset)
    ret = len(recovered)

    rnr = sum(1 for f in recovered if f["status"] == "Valid")
    corrupted = sum(1 for f in recovered if f["status"] == "Corrupted")
    fp = sum(1 for f in recovered if f["status"] == "False Positive")

    carving_rate = round((rnr / rel * 100) if rel else 0, 2)

    denom = rnr + fp + corrupted
    accuracy = round((rnr / denom * 100) if denom else 0, 2)

    per_type = {}

    for ft in FILE_TYPES_RAW:

        # Count original files of this type
        orig = sum(
            1 for v in dataset.values()
            if v["type"] == ft
        )

        # Remove duplicates using SHA-256
        valid_hashes = set(
            f["sha"] for f in recovered
            if f["status"] == "Valid"
            and f["type"] == ft
        )

        valid = len(valid_hashes)

        rate = round((valid / orig * 100) if orig else 0, 2)

        per_type[ft] = {
            "original": orig,
            "valid": valid,
            "rate": rate
        }

    return {
        "rel": rel,
        "ret": ret,
        "rnr": rnr,
        "corrupted": corrupted,
        "fp": fp,
        "carving_rate": carving_rate,
        "accuracy": accuracy,
        "per_type": per_type,
        "proc_time": proc_time
    }

# ─────────────────────────────────────────────
# READ EXISTING SUMMARY (use real data from Excel)
# ─────────────────────────────────────────────
def read_existing_summary(wb):
    """Read the pre-filled summary sheet to get accurate metrics including processing times."""
    ws = wb[SUMMARY_SHEET]
    metrics = {}
    tool_map = {"Foremost": "Foremost", "PhotoRec": "PhotoRec", "Scalpel": "Scalpel"}
    scen_map = {
        "Scenario 1 (Normal)": "Scenario 1 (Normal)",
        "Scenario 2 (Fragmented)": "Scenario 2 (Fragmented)"
    }
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        tool = str(row[0]).strip()
        scen = str(row[1]).strip() if row[1] else ""
        if tool not in tool_map or scen not in scen_map:
            continue
        rel = row[2] or 0
        ret = row[3] or 0
        rnr = row[4] or 0
        corrupted = row[5] or 0
        fp = row[6] or 0
        proc_time = str(row[7]) if row[7] else ""
        carving_rate = row[8] or 0
        accuracy = row[9] or 0
        metrics[(tool, scen)] = {
            "rel": rel, "ret": ret, "rnr": rnr,
            "corrupted": corrupted, "fp": fp,
            "carving_rate": carving_rate, "accuracy": accuracy,
            "proc_time": proc_time
        }
    return metrics


def read_existing_per_filetype(wb):
    """Read Sheet 9 per-file-type data."""
    ws = wb[FILETYPE_SHEET]
    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        ft = str(row[0]).strip().lower()
        data[ft] = {
            "Foremost_S1": row[1] or 0,
            "Foremost_S2": row[2] or 0,
            "PhotoRec_S1": row[3] or 0,
            "PhotoRec_S2": row[4] or 0,
            "Scalpel_S1": row[5] or 0,
            "Scalpel_S2": row[6] or 0,
        }
    return data


# ─────────────────────────────────────────────
# CHARTS - Section C (Results)
# Chart 1: Carving Rate Comparison
# Chart 2: Accuracy Comparison
# Chart 3: False Positives Comparison
# Chart 4: Per File Type Success Rate - Scenario 1
# Chart 5: Per File Type Success Rate - Scenario 2
# ─────────────────────────────────────────────

def chart_01_carving_rate(metrics):
    fig, ax = plt.subplots(figsize=(9, 5))
    x = np.arange(len(TOOLS))
    w = 0.35
    s1 = [metrics.get((t, SCENARIOS[0]), {}).get("carving_rate", 0) for t in TOOLS]
    s2 = [metrics.get((t, SCENARIOS[1]), {}).get("carving_rate", 0) for t in TOOLS]

    bars1 = ax.bar(x - w/2, s1, w, label="Scenario 1 (Normal)", color=SCENARIO_COLORS[0], edgecolor="white")
    bars2 = ax.bar(x + w/2, s2, w, label="Scenario 2 (Fragmented)", color=SCENARIO_COLORS[1], edgecolor="white")

    for bar in bars1 + bars2:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, h + 1, f"{h:.1f}%", ha="center", va="bottom", fontsize=9)

    ax.set_xticks(x)
    ax.set_xticklabels(TOOLS, fontsize=11)
    ax.set_ylabel("Carving Rate (%)", fontsize=11)
    ax.set_title("Chart 1 — Carving Rate Comparison", fontsize=13, fontweight="bold", pad=12)
    ax.set_ylim(0, 115)
    ax.legend(fontsize=10)
    ax.yaxis.grid(True, alpha=0.3)
    ax.set_axisbelow(True)
    fig.tight_layout()
    save_chart(fig, "01_carving_rate.png")


def chart_02_accuracy(metrics):
    fig, ax = plt.subplots(figsize=(9, 5))
    x = np.arange(len(TOOLS))
    w = 0.35
    s1 = [metrics.get((t, SCENARIOS[0]), {}).get("accuracy", 0) for t in TOOLS]
    s2 = [metrics.get((t, SCENARIOS[1]), {}).get("accuracy", 0) for t in TOOLS]

    bars1 = ax.bar(x - w/2, s1, w, label="Scenario 1 (Normal)", color=SCENARIO_COLORS[0], edgecolor="white")
    bars2 = ax.bar(x + w/2, s2, w, label="Scenario 2 (Fragmented)", color=SCENARIO_COLORS[1], edgecolor="white")

    for bar in bars1 + bars2:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, h + 0.3, f"{h:.2f}%", ha="center", va="bottom", fontsize=8)

    ax.set_xticks(x)
    ax.set_xticklabels(TOOLS, fontsize=11)
    ax.set_ylabel("Accuracy (%)", fontsize=11)
    ax.set_title("Chart 2 — Accuracy Comparison", fontsize=13, fontweight="bold", pad=12)
    ax.set_ylim(0, 42)
    ax.legend(fontsize=10)
    ax.yaxis.grid(True, alpha=0.3)
    ax.set_axisbelow(True)
    fig.tight_layout()
    save_chart(fig, "02_accuracy.png")


def chart_03_false_positives(metrics):
    fig, ax = plt.subplots(figsize=(9, 5))
    x = np.arange(len(TOOLS))
    w = 0.35
    s1 = [metrics.get((t, SCENARIOS[0]), {}).get("fp", 0) for t in TOOLS]
    s2 = [metrics.get((t, SCENARIOS[1]), {}).get("fp", 0) for t in TOOLS]

    bars1 = ax.bar(x - w/2, s1, w, label="Scenario 1 (Normal)", color=SCENARIO_COLORS[0], edgecolor="white")
    bars2 = ax.bar(x + w/2, s2, w, label="Scenario 2 (Fragmented)", color=SCENARIO_COLORS[1], edgecolor="white")

    for bar in bars1 + bars2:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, h + 5, str(int(h)), ha="center", va="bottom", fontsize=9)

    ax.set_xticks(x)
    ax.set_xticklabels(TOOLS, fontsize=11)
    ax.set_ylabel("Number of False Positives", fontsize=11)
    ax.set_title("Chart 3 — False Positives Comparison", fontsize=13, fontweight="bold", pad=12)
    ax.legend(fontsize=10)
    ax.yaxis.grid(True, alpha=0.3)
    ax.set_axisbelow(True)
    fig.tight_layout()
    save_chart(fig, "03_false_positives.png")


def chart_04_per_filetype_s1(ft_data):
    fig, ax = plt.subplots(figsize=(10, 5))
    x = np.arange(len(FILE_TYPES_RAW))
    w = 0.25

    fm = [ft_data.get(f, {}).get("Foremost_S1", 0) * 100 for f in FILE_TYPES_RAW]
    ph = [min(ft_data.get(f, {}).get("PhotoRec_S1", 0) * 100, 100) for f in FILE_TYPES_RAW]
    sc = [min(ft_data.get(f, {}).get("Scalpel_S1", 0) * 100, 100) for f in FILE_TYPES_RAW]

    ax.bar(x - w, fm, w, label="Foremost", color=TOOL_COLORS["Foremost"], edgecolor="white")
    ax.bar(x,     ph, w, label="PhotoRec", color=TOOL_COLORS["PhotoRec"], edgecolor="white")
    ax.bar(x + w, sc, w, label="Scalpel",  color=TOOL_COLORS["Scalpel"],  edgecolor="white")

    ax.set_xticks(x)
    ax.set_xticklabels(FILE_TYPES_LABEL, fontsize=11)
    ax.set_ylabel("Recovery Rate (%)", fontsize=11)
    ax.set_title("Chart 4 — Per File Type Success Rate - Scenario 1 (Normal)", fontsize=13, fontweight="bold", pad=12)
    ax.set_ylim(0, 115)
    ax.legend(fontsize=10)
    ax.yaxis.grid(True, alpha=0.3)
    ax.set_axisbelow(True)
    fig.tight_layout()
    save_chart(fig, "04_per_filetype_s1.png")


def chart_05_per_filetype_s2(ft_data):
    fig, ax = plt.subplots(figsize=(10, 5))
    x = np.arange(len(FILE_TYPES_RAW))
    w = 0.25

    fm = [ft_data.get(f, {}).get("Foremost_S2", 0) * 100 for f in FILE_TYPES_RAW]
    ph = [min(ft_data.get(f, {}).get("PhotoRec_S2", 0) * 100, 100) for f in FILE_TYPES_RAW]
    sc = [ft_data.get(f, {}).get("Scalpel_S2", 0) * 100 for f in FILE_TYPES_RAW]

    ax.bar(x - w, fm, w, label="Foremost", color=TOOL_COLORS["Foremost"], edgecolor="white")
    ax.bar(x,     ph, w, label="PhotoRec", color=TOOL_COLORS["PhotoRec"], edgecolor="white")
    ax.bar(x + w, sc, w, label="Scalpel",  color=TOOL_COLORS["Scalpel"],  edgecolor="white")

    ax.set_xticks(x)
    ax.set_xticklabels(FILE_TYPES_LABEL, fontsize=11)
    ax.set_ylabel("Recovery Rate (%)", fontsize=11)
    ax.set_title("Chart 5 — Per File Type Success Rate - Scenario 2 (Fragmented)", fontsize=13, fontweight="bold", pad=12)
    ax.set_ylim(0, 115)
    ax.legend(fontsize=10)
    ax.yaxis.grid(True, alpha=0.3)
    ax.set_axisbelow(True)
    fig.tight_layout()
    save_chart(fig, "05_per_filetype_s2.png")


# ─────────────────────────────────────────────
# CHARTS - Section D (Discussion)
# Chart 6: Fragmentation Impact (line chart)
# Chart 7: Stacked Breakdown Valid/Corrupted/FP
# ─────────────────────────────────────────────

def chart_06_fragmentation_impact(metrics):
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))

    for ax, key, title_part in zip(axes, ["carving_rate", "accuracy"], ["Carving Rate", "Accuracy"]):
        for tool in TOOLS:
            v1 = metrics.get((tool, SCENARIOS[0]), {}).get(key, 0)
            v2 = metrics.get((tool, SCENARIOS[1]), {}).get(key, 0)
            ax.plot(["Scenario 1\n(Normal)", "Scenario 2\n(Fragmented)"],
                    [v1, v2], marker="o", linewidth=2.5, markersize=8,
                    label=tool, color=TOOL_COLORS[tool])
            ax.annotate(f"{v2:.1f}%", (1, v2), textcoords="offset points",
                        xytext=(8, 0), fontsize=9, color=TOOL_COLORS[tool])

        ax.set_title(f"{title_part} — Fragmentation Impact", fontsize=12, fontweight="bold")
        ax.set_ylabel(f"{title_part} (%)", fontsize=10)
        ax.legend(fontsize=9)
        ax.yaxis.grid(True, alpha=0.3)
        ax.set_axisbelow(True)

    fig.suptitle("Chart 6 — Fragmentation Impact Analysis", fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    save_chart(fig, "06_fragmentation_impact.png")


def chart_07_stacked_breakdown(metrics):
    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    labels = TOOLS

    for ax, scen, title_suffix in zip(axes, SCENARIOS, ["Scenario 1 (Normal)", "Scenario 2 (Fragmented)"]):
        valid    = [metrics.get((t, scen), {}).get("rnr", 0) for t in TOOLS]
        corrupt  = [metrics.get((t, scen), {}).get("corrupted", 0) for t in TOOLS]
        fp_vals  = [metrics.get((t, scen), {}).get("fp", 0) for t in TOOLS]

        x = np.arange(len(labels))
        w = 0.5

        p1 = ax.bar(x, valid,   w, label="Valid",          color="#4CAF50", edgecolor="white")
        p2 = ax.bar(x, corrupt, w, bottom=valid,           label="Corrupted",      color="#FFC107", edgecolor="white")
        p3 = ax.bar(x, fp_vals, w,
                    bottom=[v+c for v,c in zip(valid, corrupt)],
                    label="False Positives", color="#F44336", edgecolor="white")

        ax.set_xticks(x)
        ax.set_xticklabels(labels, fontsize=11)
        ax.set_title(title_suffix, fontsize=11, fontweight="bold")
        ax.set_ylabel("Number of Files", fontsize=10)
        ax.legend(fontsize=9)
        ax.yaxis.grid(True, alpha=0.3)
        ax.set_axisbelow(True)

    fig.suptitle("Chart 7 — Recovery Breakdown: Valid / Corrupted / False Positives", fontsize=13, fontweight="bold", y=1.02)
    fig.tight_layout()
    save_chart(fig, "07_stacked_breakdown.png")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print(" FILE CARVING ANALYZER — UPDATED")
    print(datetime.now())
    print("=" * 60)

    # Use data_only=True to read calculated formula values
    wb = load_workbook(EXCEL_FILE, data_only=True)

    print("\n[*] Reading existing summary from Excel sheets...")
    metrics = read_existing_summary(wb)
    ft_data = read_existing_per_filetype(wb)

    for k, v in metrics.items():
        cr = v['carving_rate'] if isinstance(v['carving_rate'], (int, float)) else 0
        acc = v['accuracy'] if isinstance(v['accuracy'], (int, float)) else 0
        print(f"    {k[0]:10} | {k[1]:25} | CR={cr:.2f}% | Acc={acc:.4f}%")

    print("\n[*] Generating charts...")
    # Section C charts (Results)
    chart_01_carving_rate(metrics)
    chart_02_accuracy(metrics)
    chart_03_false_positives(metrics)
    chart_04_per_filetype_s1(ft_data)
    chart_05_per_filetype_s2(ft_data)
    # Section D charts (Discussion)
    chart_06_fragmentation_impact(metrics)
    chart_07_stacked_breakdown(metrics)

    print("\n✓ ALL 7 CHARTS GENERATED")
    print(f"  Folder: {CHARTS_DIR}/")
    for i in range(1, 8):
        prefix = f"0{i}_" if i < 10 else f"{i}_"
        names = [f for f in os.listdir(CHARTS_DIR) if f.startswith(prefix)]
        if names:
            print(f"    {names[0]}")


if __name__ == "__main__":
    main()